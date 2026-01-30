from pathlib import Path

absolute_path = Path('/Users/sivaramkjs/01_Mine/01_Coding/01_Python_Learning/hello_world_python')
# print([x for x in absolute_path.iterdir() if x.is_dir()])

relative_path = Path('../classes')
# print([x.name for x in relative_path.glob('*.py')])


with open('text.txt') as file:
    for line in file:
        print(line, end='')

# Excel file manipulation
import openpyxl as xl
from openpyxl.chart import Reference, BarChart


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)
    # print(cell.value)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Excel file charts
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
